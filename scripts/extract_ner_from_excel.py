#!/usr/bin/env python3
"""
Extract NER training data from granular annotation Excel files.

Reads phase0_extraction_note_*.xlsx files and outputs:
- ner_dataset_all.jsonl: Primary training format (one line per note with entities)
- notes.jsonl: Debug file with note texts
- spans.jsonl: Debug file with all spans
- stats.json: Label distribution and statistics
- alignment_warnings.log: Whitespace drift issues

Usage:
    python scripts/extract_ner_from_excel.py \
        --input-dir "data/granular annotations/phase0_excels" \
        --output-dir data/ml_training/granular_ner
"""

import argparse
import json
import logging
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Optional

import openpyxl

# Configure logging
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)


@dataclass
class Span:
    note_id: str
    label: str
    span_text: str
    start_char: int
    end_char: int
    hydration_status: str
    # Validation results
    is_valid: bool = True
    validation_note: str = ""


@dataclass
class Note:
    note_id: str
    note_text: str
    source_file: str


def normalize_whitespace(text: str) -> str:
    """Collapse all whitespace to single spaces."""
    return " ".join(text.split())


def validate_span(span: Span, note_text: str) -> tuple[bool, str]:
    """
    Validate that span text matches note text at given offsets.

    Returns:
        (is_valid, validation_note)
        - is_valid: True if exact match
        - validation_note: 'exact_match', 'alignment_warning', or 'alignment_error'
    """
    if span.start_char is None or span.end_char is None:
        return False, "alignment_error: missing offsets"

    try:
        extracted = note_text[span.start_char:span.end_char]
    except (IndexError, TypeError):
        return False, f"alignment_error: invalid offsets [{span.start_char}:{span.end_char}]"

    # Check exact match
    if extracted == span.span_text:
        return True, "exact_match"

    # Check normalized match (whitespace only)
    if normalize_whitespace(extracted) == normalize_whitespace(span.span_text):
        return True, "alignment_warning: whitespace mismatch"

    # Complete mismatch
    return False, f"alignment_error: extracted='{extracted[:50]}...' vs span='{span.span_text[:50]}...'"


def read_excel_file(excel_path: Path) -> tuple[Optional[Note], list[Span]]:
    """
    Read an annotation Excel file and extract note + spans.

    Returns:
        (Note or None, list of Spans)
    """
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Failed to open {excel_path}: {e}")
        return None, []

    note = None
    spans = []

    # Read Note_Text sheet
    if "Note_Text" in wb.sheetnames:
        ws = wb["Note_Text"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) >= 2:
            headers = [str(h).lower() if h else "" for h in rows[0]]
            try:
                note_id_idx = headers.index("note_id")
                note_text_idx = headers.index("note_text")

                # Find first actual data row (skip duplicate headers and empty rows)
                for data_row in rows[1:]:
                    if not data_row or not data_row[note_id_idx]:
                        continue
                    # Skip rows that look like headers (note_id column contains "note_id")
                    if str(data_row[note_id_idx]).lower() == "note_id":
                        continue
                    note = Note(
                        note_id=str(data_row[note_id_idx] or ""),
                        note_text=str(data_row[note_text_idx] or ""),
                        source_file=excel_path.name
                    )
                    break
            except (ValueError, IndexError) as e:
                logger.warning(f"Note_Text sheet missing required columns in {excel_path}: {e}")

    # Read Span_Hydrated sheet
    if "Span_Hydrated" in wb.sheetnames:
        ws = wb["Span_Hydrated"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) >= 2:
            headers = [str(h).lower() if h else "" for h in rows[0]]

            # Find required column indices
            try:
                note_id_idx = headers.index("note_id")
                label_idx = headers.index("label")
                span_text_idx = headers.index("span_text")
                start_idx = headers.index("start_char")
                end_idx = headers.index("end_char")
                status_idx = headers.index("hydration_status")
            except ValueError as e:
                logger.warning(f"Span_Hydrated sheet missing columns in {excel_path}: {e}")
                return note, spans

            # Extract spans
            for row in rows[1:]:
                if not row or len(row) <= max(note_id_idx, label_idx, span_text_idx, start_idx, end_idx, status_idx):
                    continue

                # Skip empty rows (all None) and header-like rows
                if row[note_id_idx] is None or str(row[note_id_idx]).lower() == "note_id":
                    continue

                hydration_status = str(row[status_idx] or "")

                # Filter: only keep hydrated spans
                if not hydration_status.startswith("hydrated_"):
                    continue

                # Parse start/end as integers
                try:
                    start_char = int(row[start_idx]) if row[start_idx] else None
                    end_char = int(row[end_idx]) if row[end_idx] else None
                except (ValueError, TypeError):
                    start_char = None
                    end_char = None

                if start_char is None or end_char is None:
                    continue

                span = Span(
                    note_id=str(row[note_id_idx] or ""),
                    label=str(row[label_idx] or ""),
                    span_text=str(row[span_text_idx] or ""),
                    start_char=start_char,
                    end_char=end_char,
                    hydration_status=hydration_status
                )
                spans.append(span)

    wb.close()
    return note, spans


def process_excel_files(input_dir: Path) -> tuple[list[Note], list[Span], dict]:
    """
    Process all Excel files in the input directory.

    Returns:
        (list of Notes, list of Spans, stats dict)
    """
    all_notes = []
    all_spans = []
    stats = {
        "total_files": 0,
        "successful_files": 0,
        "total_notes": 0,
        "total_spans_raw": 0,
        "total_spans_valid": 0,
        "alignment_warnings": 0,
        "alignment_errors": 0,
        "label_counts": Counter(),
        "hydration_status_counts": Counter()
    }

    excel_files = sorted(input_dir.glob("phase0_extraction_*.xlsx"))
    stats["total_files"] = len(excel_files)

    logger.info(f"Found {len(excel_files)} Excel files in {input_dir}")

    alignment_warnings = []

    for excel_path in excel_files:
        note, spans = read_excel_file(excel_path)

        if note is None:
            logger.warning(f"No note found in {excel_path.name}")
            continue

        stats["successful_files"] += 1
        stats["total_notes"] += 1
        all_notes.append(note)

        # Validate each span
        for span in spans:
            stats["total_spans_raw"] += 1
            stats["hydration_status_counts"][span.hydration_status] += 1

            is_valid, validation_note = validate_span(span, note.note_text)
            span.is_valid = is_valid
            span.validation_note = validation_note

            if "alignment_warning" in validation_note:
                stats["alignment_warnings"] += 1
                alignment_warnings.append({
                    "note_id": span.note_id,
                    "label": span.label,
                    "span_text": span.span_text[:50],
                    "start": span.start_char,
                    "end": span.end_char,
                    "issue": validation_note
                })
            elif "alignment_error" in validation_note:
                stats["alignment_errors"] += 1
                alignment_warnings.append({
                    "note_id": span.note_id,
                    "label": span.label,
                    "span_text": span.span_text[:50],
                    "start": span.start_char,
                    "end": span.end_char,
                    "issue": validation_note
                })
                continue  # Skip invalid spans

            if is_valid:
                stats["total_spans_valid"] += 1
                stats["label_counts"][span.label] += 1
                all_spans.append(span)

    stats["label_counts"] = dict(stats["label_counts"].most_common())
    stats["hydration_status_counts"] = dict(stats["hydration_status_counts"])
    stats["alignment_issues"] = alignment_warnings

    return all_notes, all_spans, stats


def write_outputs(notes: list[Note], spans: list[Span], stats: dict, output_dir: Path):
    """Write all output files."""
    output_dir.mkdir(parents=True, exist_ok=True)

    # Group spans by note_id
    spans_by_note = defaultdict(list)
    for span in spans:
        spans_by_note[span.note_id].append(span)

    # 1. Primary output: ner_dataset_all.jsonl
    ner_path = output_dir / "ner_dataset_all.jsonl"
    with open(ner_path, "w") as f:
        for note in notes:
            note_spans = spans_by_note.get(note.note_id, [])
            # Sort entities by start position
            entities = sorted([
                {
                    "start": s.start_char,
                    "end": s.end_char,
                    "label": s.label,
                    "text": s.span_text
                }
                for s in note_spans
            ], key=lambda x: x["start"])

            record = {
                "id": note.note_id,
                "text": note.note_text,
                "entities": entities
            }
            f.write(json.dumps(record) + "\n")
    logger.info(f"Wrote {len(notes)} records to {ner_path}")

    # 2. Debug: notes.jsonl
    notes_path = output_dir / "notes.jsonl"
    with open(notes_path, "w") as f:
        for note in notes:
            f.write(json.dumps({"note_id": note.note_id, "note_text": note.note_text}) + "\n")
    logger.info(f"Wrote {len(notes)} notes to {notes_path}")

    # 3. Debug: spans.jsonl
    spans_path = output_dir / "spans.jsonl"
    with open(spans_path, "w") as f:
        for span in spans:
            f.write(json.dumps({
                "note_id": span.note_id,
                "label": span.label,
                "span_text": span.span_text,
                "start_char": span.start_char,
                "end_char": span.end_char,
                "hydration_status": span.hydration_status
            }) + "\n")
    logger.info(f"Wrote {len(spans)} spans to {spans_path}")

    # 4. Stats
    stats_path = output_dir / "stats.json"
    # Remove alignment_issues from stats for cleaner output
    alignment_issues = stats.pop("alignment_issues", [])
    with open(stats_path, "w") as f:
        json.dump(stats, f, indent=2)
    logger.info(f"Wrote stats to {stats_path}")

    # 5. Alignment warnings log
    if alignment_issues:
        warnings_path = output_dir / "alignment_warnings.log"
        with open(warnings_path, "w") as f:
            for issue in alignment_issues:
                f.write(json.dumps(issue) + "\n")
        logger.warning(f"Wrote {len(alignment_issues)} alignment issues to {warnings_path}")


def main():
    parser = argparse.ArgumentParser(description="Extract NER training data from Excel files")
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("data/granular annotations/phase0_excels"),
        help="Directory containing phase0_extraction_*.xlsx files"
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("data/ml_training/granular_ner"),
        help="Directory for output files"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Print stats without writing files"
    )
    args = parser.parse_args()

    logger.info(f"Input directory: {args.input_dir}")
    logger.info(f"Output directory: {args.output_dir}")

    if not args.input_dir.exists():
        logger.error(f"Input directory does not exist: {args.input_dir}")
        return 1

    notes, spans, stats = process_excel_files(args.input_dir)

    # Print summary
    print("\n=== Extraction Summary ===")
    print(f"Files processed: {stats['successful_files']}/{stats['total_files']}")
    print(f"Notes extracted: {stats['total_notes']}")
    print(f"Spans (raw): {stats['total_spans_raw']}")
    print(f"Spans (valid): {stats['total_spans_valid']}")
    print(f"Alignment warnings: {stats['alignment_warnings']}")
    print(f"Alignment errors: {stats['alignment_errors']}")
    print(f"\nLabel distribution (top 10):")
    for label, count in list(stats['label_counts'].items())[:10]:
        print(f"  {label}: {count}")

    if args.dry_run:
        print("\n[Dry run - no files written]")
        return 0

    write_outputs(notes, spans, stats, args.output_dir)

    print(f"\nOutput written to: {args.output_dir}")
    return 0


if __name__ == "__main__":
    exit(main())
